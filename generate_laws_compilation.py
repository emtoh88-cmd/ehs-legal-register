#!/usr/bin/env python3
"""
generate_laws_compilation.py — export register.json as a Laws Compilation
workbook, in the same shape (Index / Update History / Contents / SSCP sheets)
as the paid EHS compliance subscription used to provide, for personal
offline reference.

This is a straight dump of your own register.json -- no data is fetched.
Run sso_refresh.py first if you want the export to reflect the latest
checked instruments.

Usage:
  python generate_laws_compilation.py                # writes reports/laws-compilation-YYYY-MM-DD.xlsx
  python generate_laws_compilation.py --out FILE.xlsx
"""

import argparse, json, os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REGISTER = "register.json"
REPORTS_DIR = "reports"

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True)


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    for i, h in enumerate(headers, 1):
        col = get_column_letter(i)
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[col].width = min(width + 2, 80)
    ws.freeze_panes = "A2"


def build_xlsx(reg, out_path):
    wb = Workbook()

    ws = wb.active
    ws.title = "Contents"
    write_sheet(ws, ["Category", "Sub-category", "Citation", "Reference / SSO ID"],
                [[c.get("cat", ""), c.get("sub") or "", c.get("citation", ""), c.get("ref", "")]
                 for c in reg["contents"]])

    ws = wb.create_sheet("Index")
    write_sheet(ws, ["Citation", "Reference", "SSO path", "Applicability", "Last checked"],
                [[e.get("citation", ""), e.get("ref", ""), e.get("sso", ""),
                  e.get("applicability", ""), e.get("checked", "")]
                 for e in reg["index"]])

    ws = wb.create_sheet("SSCP")
    write_sheet(ws, ["Group", "Title", "Agency", "Applicability", "Summary"],
                [[s.get("group", ""), s.get("title", ""), s.get("agency", ""),
                  s.get("applicability", ""), s.get("summary", "")]
                 for s in reg["sscp"]])

    ws = wb.create_sheet("Update History")

    def hist_sort_key(h):
        try:
            return (h.get("wef") or "", int(h.get("no", 0)))
        except ValueError:
            return (h.get("wef") or "", 0)

    history_rows = sorted(reg["history"], key=hist_sort_key, reverse=True)
    write_sheet(ws, ["Quarter", "S/No", "Category", "Title", "Wef", "Vide"],
                [[h.get("quarter", ""), h.get("no", ""), h.get("category", ""),
                  h.get("title", ""), h.get("wef", ""), h.get("vide", "")]
                 for h in history_rows])

    meta = reg.get("meta", {})
    ws = wb.create_sheet("About")
    ws.append(["Generated", date.today().isoformat()])
    ws.append(["Source snapshot", meta.get("source_snapshot", "")])
    ws.append(["Last checked", meta.get("last_checked", "")])
    ws.append(["Instruments (Index)", len(reg["index"])])
    ws.append(["Standards / COPs (SSCP)", len(reg["sscp"])])
    ws.append(["Recorded changes (History)", len(reg["history"])])
    ws.append([])
    ws.append([meta.get("notice", "")])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--out", help="output .xlsx path (default: reports/laws-compilation-YYYY-MM-DD.xlsx)")
    args = p.parse_args()

    if not os.path.exists(REGISTER):
        print(f"{REGISTER} not found"); return 1
    reg = json.load(open(REGISTER, encoding="utf-8"))

    os.makedirs(REPORTS_DIR, exist_ok=True)
    out_path = args.out or os.path.join(REPORTS_DIR, f"laws-compilation-{date.today():%Y-%m-%d}.xlsx")

    build_xlsx(reg, out_path)
    print(f"wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
